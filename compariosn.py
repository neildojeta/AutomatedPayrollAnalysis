import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Load the Excel files
file_previous = "DONE-VDP_DIV5_1007_1020_FINAL.xlsm"
file_latest = "VDP_DIV5_1021_1103_FINAL.xlsm"

# Load the relevant sheets
sheet_pr_previous = pd.read_excel(file_previous, sheet_name="PR DATE")
pd.set_option('display.max_rows', None)

# Check the number of rows in the sheet
num_rows = sheet_pr_previous.shape[0]
print(f"The sheet contains {num_rows} rows.")

# Define a function to clean the formatted string and convert it to a float
def clean_currency(value):
    if isinstance(value, str):
        # Remove dollar signs, commas, and spaces before converting to float
        value = value.replace('$', '').replace(',', '').strip()
        try:
            return round(float(value), 2) if value else None
        except ValueError:
            return None
    return value

# Apply the clean_currency function to the DataFrame
cleaned_values = sheet_pr_previous.applymap(clean_currency)

# Remove completely empty rows (rows with all NaN or None values)
# cleaned_values = cleaned_values.dropna(how='all')

# Print out the entire cleaned DataFrame
print("Cleaned values preview:")
print(cleaned_values)

# Define the value to search for
search_value = 93346.83

# Search for the cleaned value in the DataFrame
matching_coords = cleaned_values[cleaned_values == search_value]

# Check if any match is found and get coordinates
if matching_coords.any().any():
    row, col = matching_coords.stack().index[0]
    print(f"Value found at row {row}, column {col}")
else:
    print("Value not found.")

sheet_pr_latest = pd.read_excel(file_latest, sheet_name="PR DATE")
sheet_hours_previous = pd.read_excel(file_previous, sheet_name="Hours_Working")
sheet_hours_latest = pd.read_excel(file_latest, sheet_name="Hours_Working")

# Function to calculate totals from "Hours_Working"
def calculate_totals(hours_sheet, pr_sheet, client=None):
    totals = {
        "TRIPS": hours_sheet["TRIPS"].sum(),
        "HOURS": hours_sheet["SERVICE HOURS OPERATED"].sum(),
        "OPERATORS": hours_sheet["OPERATOR NAME"].nunique(),
        "DAYS": hours_sheet["Date"].nunique(),
        "AMOUNT" : pr_sheet.iloc[79, 14] if len(pr_sheet) > 79 else 0
    }
    
    # Logging to check if AMOUNT is properly added
    print(f"AMOUNT for client {client}: {totals.get('AMOUNT', 'Not Found')}")  # Debugging line
    
    return totals

# Function to calculate totals from "Hours_Working"
def calculate_client_totals(hours_sheet, pr_sheet, client):
    if client == "LAVTA":
        totals = {
        "TRIPS": hours_sheet["TRIPS"].sum(),
        "HOURS": hours_sheet["SERVICE HOURS OPERATED"].sum(),
        "OPERATORS": hours_sheet["OPERATOR NAME"].nunique(),
        "DAYS": hours_sheet["Date"].nunique(),
        "AMOUNT" : pr_sheet.iloc[76, 14] if len(pr_sheet) > 76 else 0
        }
    else:
        totals = {
        "TRIPS": hours_sheet["TRIPS"].sum(),
        "HOURS": hours_sheet["SERVICE HOURS OPERATED"].sum(),
        "OPERATORS": hours_sheet["OPERATOR NAME"].nunique(),
        "DAYS": hours_sheet["Date"].nunique(),
        "AMOUNT" : pr_sheet.iloc[27, 14] if len(pr_sheet) > 27 else 0
        }
    # Logging to check if AMOUNT is properly added
    print(f"AMOUNT for client {client}: {totals.get('AMOUNT', 'Not Found')}")  # Debugging line
    
    return totals

# Function to compare operators and extract partner information from the "Hours_Working" sheet
def compare_operators(sheet_previous, sheet_latest):
    operators_previous = set(sheet_previous[["OPERATOR NAME", "PARTNER"]].dropna().itertuples(index=False, name=None))
    operators_latest = set(sheet_latest[["OPERATOR NAME", "PARTNER"]].dropna().itertuples(index=False, name=None))

    added = operators_latest - operators_previous
    removed = operators_previous - operators_latest

    added_list = [{"Operator Name": op, "Partner": partner} for op, partner in added]
    removed_list = [{"Operator Name": op, "Partner": partner} for op, partner in removed]

    return {"Added": added_list, "Removed": removed_list}

# Function to compare trips and service hours per partner
def compare_trips_and_hours(sheet_previous, sheet_latest):
    grouped_previous = sheet_previous.groupby("PARTNER")[["TRIPS", "SERVICE HOURS OPERATED"]].sum()
    grouped_latest = sheet_latest.groupby("PARTNER")[["TRIPS", "SERVICE HOURS OPERATED"]].sum()

    comparison = grouped_previous.join(grouped_latest, how="outer", lsuffix="_PREVIOUS", rsuffix="_LATEST").fillna(0)
    comparison["TRIPS_CHANGE"] = comparison["TRIPS_LATEST"] - comparison["TRIPS_PREVIOUS"]
    comparison["HOURS_CHANGE"] = comparison["SERVICE HOURS OPERATED_LATEST"] - comparison["SERVICE HOURS OPERATED_PREVIOUS"]

    trips_comparison = comparison[["TRIPS_PREVIOUS", "TRIPS_LATEST", "TRIPS_CHANGE"]].reset_index()
    trips_comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

    hours_comparison = comparison[["SERVICE HOURS OPERATED_PREVIOUS", "SERVICE HOURS OPERATED_LATEST", "HOURS_CHANGE"]].reset_index()
    hours_comparison.columns = ["PARTNER", "PREVIOUS", "LATEST", "CHANGE"]

    return trips_comparison, hours_comparison

# Apply formatting to the Excel sheets
def apply_formatting(sheet_name, wb):
    ws = wb[sheet_name]
    # Format headers: Blue background, white font
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")  # White text on blue background
        cell.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Add borders and conditional formatting
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

            # Apply conditional formatting to "Change"/"CHANGE" column
            if ws[1][cell.column - 1].value.lower() == "change":
                if isinstance(cell.value, (int, float)):  # Check if numeric
                    if cell.value > 0:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                        cell.font = Font(color="006100")
                    elif cell.value < 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                        cell.font = Font(color="9C0006")
                elif isinstance(cell.value, str):  # Handle "Increased", "Decreased", "Added", and "Removed"
                    if cell.value.lower() in ["increased", "added"]:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                        cell.font = Font(color="006100")
                    elif cell.value.lower() in ["decreased", "removed"]:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                        cell.font = Font(color="9C0006")

# Main comparison and file saving logic
output_folder = "Compared Results"
os.makedirs(output_folder, exist_ok=True)

# 1. Process the data without any filtering (full comparison)
# Recalculate totals for both previous and latest
totals_previous = calculate_totals(sheet_hours_previous, sheet_pr_previous)
totals_latest = calculate_totals(sheet_hours_latest, sheet_pr_latest)

# Calculate differences and changes
differences = {
    "TRIPS": totals_latest["TRIPS"] - totals_previous["TRIPS"],
    "HOURS": totals_latest["HOURS"] - totals_previous["HOURS"],
    "OPERATORS": totals_latest["OPERATORS"] - totals_previous["OPERATORS"],
    "DAYS": totals_latest["DAYS"] - totals_previous["DAYS"],
    "AMOUNT": totals_latest["AMOUNT"] - totals_previous["AMOUNT"],
}
changes = {key: "Increased" if diff > 0 else "Decreased" if diff < 0 else "No Change"
            for key, diff in differences.items()}

# Create summary DataFrame
summary_table = {
    "Metric": ["TRIPS", "HOURS", "OPERATORS", "DAYS", "AMOUNT"],
    "Previous": [totals_previous[key] for key in totals_previous],
    "Latest": [totals_latest[key] for key in totals_latest],
    "Difference": [differences[key] for key in differences],
    "Change": [changes[key] for key in changes],
}
summary_df = pd.DataFrame(summary_table)

# Compare operators
operator_changes = compare_operators(sheet_hours_previous, sheet_hours_latest)
added_df = pd.DataFrame(operator_changes["Added"])
removed_df = pd.DataFrame(operator_changes["Removed"])
added_df["Change"] = "Added"
removed_df["Change"] = "Removed"
operator_changes_df = pd.concat([added_df, removed_df], ignore_index=True)

# Compare trips and hours
trips_comparison_df, hours_comparison_df = compare_trips_and_hours(sheet_hours_previous, sheet_hours_latest)

# Save the full comparison results
full_comparison_file = os.path.join(output_folder, "Full_Comparison.xlsx")
with pd.ExcelWriter(full_comparison_file, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    operator_changes_df.to_excel(writer, sheet_name="Operator Changes", index=False)
    trips_comparison_df.to_excel(writer, sheet_name="Trips Comparison", index=False)
    hours_comparison_df.to_excel(writer, sheet_name="Hours Comparison", index=False)

# Apply formatting to the full comparison file
wb_full = load_workbook(full_comparison_file)
for sheet in ["Summary", "Operator Changes", "Trips Comparison", "Hours Comparison"]:
    apply_formatting(sheet, wb_full)
wb_full.save(full_comparison_file)

# 2. Process data for each client
unique_clients = sheet_hours_latest["CLIENT"].dropna().unique()

for client in unique_clients:
    # Filter by client for both previous and latest sheets
    sheet_previous_client = sheet_hours_previous[sheet_hours_previous["CLIENT"] == client]
    sheet_latest_client = sheet_hours_latest[sheet_hours_latest["CLIENT"] == client]

    # Recalculate totals for client
    totals_previous_client = calculate_client_totals(sheet_previous_client, sheet_pr_previous, client)
    totals_latest_client = calculate_client_totals(sheet_latest_client, sheet_pr_latest, client)

    # Calculate differences and changes for the client
    differences_client = {
        "TRIPS": totals_latest_client["TRIPS"] - totals_previous_client["TRIPS"],
        "HOURS": totals_latest_client["HOURS"] - totals_previous_client["HOURS"],
        "OPERATORS": totals_latest_client["OPERATORS"] - totals_previous_client["OPERATORS"],
        "DAYS": totals_latest_client["DAYS"] - totals_previous_client["DAYS"],
        "AMOUNT": totals_latest_client["AMOUNT"] - totals_previous_client["AMOUNT"],
    }
    changes_client = {key: "Increased" if diff > 0 else "Decreased" if diff < 0 else "No Change"
                      for key, diff in differences_client.items()}

    # Create summary DataFrame for the client
    summary_table_client = {
        "Metric": ["TRIPS", "HOURS", "OPERATORS", "DAYS", "AMOUNT"],
        "Previous": [totals_previous_client[key] for key in totals_previous_client],
        "Latest": [totals_latest_client[key] for key in totals_latest_client],
        "Difference": [differences_client[key] for key in differences_client],
        "Change": [changes_client[key] for key in changes_client],
    }
    summary_df_client = pd.DataFrame(summary_table_client)

    # Compare operators for the client
    operator_changes_client = compare_operators(sheet_previous_client, sheet_latest_client)
    added_df_client = pd.DataFrame(operator_changes_client["Added"])
    removed_df_client = pd.DataFrame(operator_changes_client["Removed"])
    added_df_client["Change"] = "Added"
    removed_df_client["Change"] = "Removed"
    operator_changes_df_client = pd.concat([added_df_client, removed_df_client], ignore_index=True)

    # Compare trips and hours for the client
    trips_comparison_df_client, hours_comparison_df_client = compare_trips_and_hours(sheet_previous_client, sheet_latest_client)

    # Save output for the client
    client_output_file = os.path.join(output_folder, f"{client}_Comparison.xlsx")
    with pd.ExcelWriter(client_output_file, engine="openpyxl") as writer:
        summary_df_client.to_excel(writer, sheet_name="Summary", index=False)
        operator_changes_df_client.to_excel(writer, sheet_name="Operator Changes", index=False)
        trips_comparison_df_client.to_excel(writer, sheet_name="Trips Comparison", index=False)
        hours_comparison_df_client.to_excel(writer, sheet_name="Hours Comparison", index=False)

    # Apply formatting to the client's output
    wb_client = load_workbook(client_output_file)
    for sheet in ["Summary", "Operator Changes", "Trips Comparison", "Hours Comparison"]:
        apply_formatting(sheet, wb_client)
    wb_client.save(client_output_file)

print("Full comparison and filtered comparisons for each client are complete!")

